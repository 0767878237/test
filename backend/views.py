from django.shortcuts import render, redirect
from django.http import HttpResponse
from backend.models import CrawlTask, BusinessData, Login
from sqlalchemy.orm import sessionmaker
# from .forms import CrawlTaskForm
from django.core.files.storage import FileSystemStorage
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
import os
import subprocess
from django.contrib import messages
from django.db import connection
from crawl_runner import create_and_run_task
from django.core.exceptions import ObjectDoesNotExist
from django.views.decorators.cache import never_cache
from functools import wraps
from django.contrib.auth.models import User
from django.core.mail import send_mail
from django.contrib.auth.tokens import PasswordResetTokenGenerator
# block cache and protect views
def login_required_no_cache(view_func):
    @wraps(view_func)
    @never_cache
    def _wrapped_view(request, *args, **kwargs):
        if not request.session.get('user_email'):
            return redirect('sign_in')
        return view_func(request, *args, **kwargs)
    return _wrapped_view
# check email in home page and redirect to sign in page if not logged in
def require_login(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.session.get('user_email'):
            return redirect('sign_in')
        return view_func(request, *args, **kwargs)
    return wrapper

# show main page
@login_required_no_cache
@require_login
def main(request):
    user_email = request.session.get('user_email')
    tasks = CrawlTask.objects.all().order_by('-created_at')
    return render(request, 'crawler_app/main.html', {'tasks': tasks, 'user_email': user_email})
def view(request):
    tasks = CrawlTask.objects.all().order_by('-created_at')
    return render(request, 'crawler_app/view.html', {'tasks': tasks})

# show dashboard page
def index(request):
    tasks = CrawlTask.objects.all().order_by('-created_at')
    return render(request, 'crawler_app/index.html') 

# create new tasks
def create_task(request):
    if request.method == 'POST':
        url_filter = request.POST.get('url_filter')  # Lấy giá trị từ form
        if url_filter:
            # Lưu dữ liệu vào bảng CrawlTask
            CrawlTask.objects.create(url_filter=url_filter)
            #create_and_run_task(url_filter)
            return redirect('home')  # Sau khi tạo task xong, redirect về trang chính
    return render(request, 'crawler_app/index.html')  # Nếu không phải POST, render lại trang

# delete tasks
def delete_crawltask(request, task_id):
    if request.method == 'POST':
        try:
            task = CrawlTask.objects.get(id=task_id)
            task.delete()
            reset_crawltask_sequence()  # đặt id = max + 1
            messages.success(request, 'Đã xoá task thành công.')
        except CrawlTask.DoesNotExist:
            messages.warning(request, 'Task không tồn tại hoặc đã bị xoá.')
    return redirect('home')

# đặt id = max + 1
def reset_crawltask_sequence():
    with connection.cursor() as cursor:
        cursor.execute("SELECT MAX(id) FROM backend_crawltask;")
        row = cursor.fetchone()
        max_id = row[0] if row[0] else 0

        # Cập nhật lại giá trị trong sqlite_sequence để đảm bảo giá trị ID kế tiếp là max_id + 1
        cursor.execute("UPDATE sqlite_sequence SET seq = ? WHERE name = ?", (max_id, 'backend_crawltask'))
        
# search tasks
def search_task(request):
    task_id = request.GET.get('task_id')
    task_id = int(task_id)
    try:
        task = CrawlTask.objects.filter(id=task_id)
        return render(request, 'crawler_app/main.html', {'tasks': task})
    except ObjectDoesNotExist:
        return render(request, 'crawler_app/main.html', {'error': f'Không tìm thấy task với ID {task_id}'})

import sys
import threading
import logging
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
def run_task_in_background(task_id, url_filter):
    logger.info(f"Starting task {task_id} in background...")

    try:
        # Lấy đường dẫn tuyệt đối của script crawl_runner.py
        script_path = os.path.abspath("crawl_runner.py")
        
        # Sử dụng sys.executable để lấy Python trong môi trường ảo
        python_path = sys.executable  # Đây sẽ là Python trong môi trường ảo
        
        # In ra thông tin về môi trường ảo và script đang chạy
        logger.info(f"Python executable: {python_path}")
        logger.info(f"Running script: {script_path}")

        # Chạy subprocess với Python trong môi trường ảo và gọi hàm create_and_run_task trong script
        process = subprocess.Popen([python_path, script_path, str(task_id), url_filter], stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        
        # Đọc kết quả từ subprocess
        stdout, stderr = process.communicate()

        # Kiểm tra và in ra output
        if stdout:
            logger.info(f"Output: {stdout.decode()}")
        if stderr:
            logger.error(f"Error: {stderr.decode()}")

        logger.info(f"Task {task_id} started.")
    except Exception as e:
        logger.error(f"Error running subprocess: {e}")


def action_task(request, task_id):
    if request.method == 'GET':
        print("Hello")
        task = CrawlTask.objects.get(id=task_id)
        # Khởi chạy task trong background bằng thread
        threading.Thread(target=run_task_in_background, args=(task_id, task.url_filter)).start()
        return redirect('home')
    return render(request, 'crawler_app/index.html')


import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

#Export data
def export_data(request, task_id):
    task = CrawlTask.objects.get(id=task_id)
    lst_data = BusinessData.objects.filter(task_id=task_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Exported_data_task_{task_id}"

    # STYLE
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # HEADERS
    headers = ["Company Name", "Phone", "Address", "Category", "Website", "Email"]
    ws.append(headers)
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_num)].width = 40  # Tăng chiều rộng

    # DATA ROWS
    for row_idx, data in enumerate(lst_data, start=2):
        values = [data.name, data.phone, data.address, data.category, data.website, data.email]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # RETURN RESPONSE
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"export_task_{task_id}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

#========================= SIGN IN ============================

def sign_in(request):
    return render(request, 'crawler_app/sign_in.html')

def login(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        passw = request.POST.get('passw')
        try:
            check = Login.objects.get(email=email)
            if check.password == passw:
                # ✅ Lưu thông tin người dùng vào session
                request.session['user_id'] = check.id
                request.session['user_email'] = check.email
                return redirect('home')
            else:
                return render(request, 'crawler_app/sign_in.html', {'error': 'Incorrect password!'})
        except Login.DoesNotExist:
            return render(request, 'crawler_app/sign_in.html', {'error': 'Email not found!'})

    return render(request, 'crawler_app/sign_in.html')


#========================= SIGN UP ============================
def sign_up(request):
    return render(request, 'crawler_app/sign_up.html')

def register_user(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        passw = request.POST.get('passw')
        confirm = request.POST.get('c_passw')
        if email == "" or passw == "":
            return render(request, 'sign_up.html')
        try:
            temp = Login.objects.get(email=email)
        except Login.DoesNotExist:
            temp = None
            
        if temp is None:
            if(passw != confirm):
                return render(request, 'sign_up.html', {'error': 'Password do not match!'})
            
            Login.objects.create(email=email, password=passw)
            return redirect('sign_in')
        else:
            return render(request, 'crawler_app/sign_up.html', {'error': 'Email already exists!'})
        
    return render(request, 'crawler_app/sign_in.html')

#========================= LOG OUT ============================
def logout(request):
    request.session.flush()  # Xoá toàn bộ session
    return redirect('sign_in')  # Quay về trang đăng nhập

from django.contrib.auth.tokens import PasswordResetTokenGenerator
from django.contrib.auth.models import User
from django.core.mail import send_mail
from django.contrib import messages
from django.shortcuts import render, redirect
#========================= RESET PASSWORD ============================
def reset_password(request):
    return render(request, 'crawler_app/reset_password.html')

def password_reset_request(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        try:
            user = Login.objects.get(email=email)
            id = user.id
            return render(request, 'crawler_app/new_password.html', {'id': id})
        except Login.DoesNotExist:
            messages.error(request, 'Email không tồn tại.')
            return redirect('reset_password')
    return render(request, 'crawler_app/reset_password.html')
#========================= NEW PASSWORD ============================
def new_password(request):
    try:
        if request.method == 'POST':
            id = request.POST.get('user_id')
            password = request.POST.get('password')
            confirm = request.POST.get('confirm-password')
            user = Login.objects.get(id=id)
            if password == confirm:
                user.password = confirm  # Bạn nên mã hóa mật khẩu nếu dùng thật
                user.save()
                messages.success(request, 'Đã đặt lại mật khẩu thành công!')
                return redirect('sign_in')
            else:
                messages.error(request, 'Mật khẩu không khớp.')
    except Login.DoesNotExist:
        messages.error(request, 'Người dùng không tồn tại.')
    return render(request, 'crawler_app/new_password.html')
# def new_password(request):
#     if request.method == 'POST':
#         password = request.POST.get('password')
#         confirm_password = request.POST.get('confirm-password')

#         if password != confirm_password:
#             return render(request, 'crawler_app/new_password.html', {'error': 'Passwords do not match!'})
#         return redirect('sign_in')
#     return render(request, 'crawler_app/new_password.html')